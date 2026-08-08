#!/usr/bin/env python3
"""
Build the six data-pack SKUs sold on /data-packs from the engine data files.

Every row in every file here is exported from a JSON file that one of the five
applications actually reads at runtime. Nothing is authored in this script: if a
number is not in the engine, it does not appear in a pack. That is what makes
the "same library the tool runs on" claim on the product page true rather than
a marketing line.

    Pack 1  Building Code Compliance Library   <- Blueprint Auditor  resources/rules/*.json
    Pack 2  Assembly Property Schema           <- SpecChecker        knowledge/{attribute_rules,csi_divisions}.json
    Pack 3  Estimating Rate Pack 2026          <- AI Estimator       pricing/*.json + risk/risk_rules.json
    Pack 4  Bid-Levelling Template             <- Bid Leveler        knowledge/trade_scopes.json
    Pack 5  Master Scope Requirement Checklist <- ScopeGuard         knowledge/{trade_scopes,ontario_intel,interface_matrix}.json
    Bundle  all five, unmodified, plus one index

Coverage is Ontario and the Toronto/GTA municipal layer, because that is the
only jurisdiction the engines carry data for. Packs 2, 4 and 5 are largely
standards-based (CSI MasterFormat) and travel; packs 1 and 3 are statutory and
do not. Each pack states its own coverage in its README and on its title page —
a buyer must be able to tell what they bought from the file itself.

Output lands in packs-dist/, which is gitignored: these are paid files and must
not be committed or served from public/. Upload them to private storage and set
the PACK_URL_* env vars (see lib/fulfillment.js) to switch delivery to instant.

Usage:  python3 scripts/build_data_packs.py [--engines PATH] [--out PATH]
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
import zipfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    KeepTogether,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

# ── Constants ────────────────────────────────────────────────────────────────

REPO = Path(__file__).resolve().parent.parent
DEFAULT_ENGINES = REPO.parent / "MaraponeAI-Tools"
DEFAULT_OUT = REPO / "packs-dist"
LOGO = REPO / "public" / "MARACON-WHITE.png"

BUILD_DATE = date.today().isoformat()
EDITION = "2026"

DISCLAIMER = (
    "Reference material only — not a substitute for review by a licensed "
    "professional in your jurisdiction."
)

# Pulled off the site's own palette so a pack looks like the site it came from.
INK = colors.HexColor("#12161A")       # near-black header band
CHALK = colors.HexColor("#FFFFFF")
HIVIZ = colors.HexColor("#FF6A13")     # safety orange accent
FOG = colors.HexColor("#4A5560")
PLATE = colors.HexColor("#D9DEE3")
BAND = colors.HexColor("#F2F4F6")

# ── Small helpers ────────────────────────────────────────────────────────────


def load(path: Path):
    with path.open(encoding="utf-8") as fh:
        return json.load(fh)


def clean(value) -> str:
    """Flatten any JSON scalar/list into one trimmed cell of text."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, (list, tuple)):
        return "; ".join(clean(v) for v in value if v not in (None, ""))
    if isinstance(value, dict):
        return "; ".join(f"{k}: {clean(v)}" for k, v in value.items())
    return str(value).strip()


def pct(value) -> str:
    """0.13 -> '13%'. Rates are stored as fractions; buyers read percentages."""
    if value in (None, ""):
        return ""
    return f"{float(value) * 100:g}%"


# Engine keys are snake_case and heavy with acronyms, so a plain .title() turns
# statute names into "Wsib" and "Ccdc2". These are the ones that appear as
# headings in the packs.
_ACRONYMS = {
    "wsib", "ohsa", "hst", "gst", "pst", "vat", "obc", "ofc", "aoda", "csi",
    "esa", "ei", "cpp", "eht", "rfi", "dc", "dcs", "ccdc2", "cca1", "cca",
    "ccdc", "stc", "shgc", "merv", "aic", "gc", "ul", "ulc", "csa", "ansi",
    "gfa", "lm", "cor", "id", "pdf", "csv", "json",
}


def humanise(key) -> str:
    """'wsib_rate_group' -> 'WSIB Rate Group'."""
    words = str(key).replace("_", " ").split()
    return " ".join(w.upper() if w.lower() in _ACRONYMS else w.capitalize()
                    for w in words)


def section_of(clause: str) -> str:
    """'OBC 2025 Article 9.8.4.1(1)(a)' -> '9.8.4.1'."""
    match = re.search(r"\d+(?:\.\d+)+", clause or "")
    return match.group(0) if match else ""


def write_csv(path: Path, header: list[str], rows: list[list]) -> Path:
    """UTF-8 with BOM so Excel opens accented text correctly on both platforms."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        writer.writerow(header)
        writer.writerows([[clean(cell) for cell in row] for row in rows])
    return path


def write_text(path: Path, body: str) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(body.strip() + "\n", encoding="utf-8")
    return path


# ── The one branded PDF template, built once and reused by every pack ─────────


class PackDoc(BaseDocTemplate):
    """Dark header band with the wordmark, footer with disclaimer and page number.

    The disclaimer is drawn by the page template rather than added as flowable
    text, which is what makes "every jurisdiction file, no exceptions" hold
    automatically — it cannot be forgotten on a page.
    """

    def __init__(self, path: Path, title: str, subtitle: str, wide: bool = False):
        size = landscape(letter) if wide else letter
        path.parent.mkdir(parents=True, exist_ok=True)
        super().__init__(
            str(path),
            pagesize=size,
            leftMargin=0.7 * inch,
            rightMargin=0.7 * inch,
            topMargin=1.15 * inch,
            bottomMargin=0.95 * inch,
            title=title,
            author="Marapone Contracting",
            subject=subtitle,
            creator="Marapone data-pack build",
        )
        self.doc_title = title
        self.doc_subtitle = subtitle
        frame = Frame(
            self.leftMargin,
            self.bottomMargin,
            self.width,
            self.height,
            id="body",
            leftPadding=0,
            rightPadding=0,
            topPadding=0,
            bottomPadding=0,
        )
        self.addPageTemplates([PageTemplate(id="pack", frames=[frame], onPage=self._chrome)])

    def _chrome(self, canvas, doc):
        canvas.saveState()
        width, height = doc.pagesize

        # Header band
        band_h = 0.72 * inch
        canvas.setFillColor(INK)
        canvas.rect(0, height - band_h, width, band_h, stroke=0, fill=1)
        canvas.setFillColor(HIVIZ)
        canvas.rect(0, height - band_h - 2.5, width, 2.5, stroke=0, fill=1)

        if LOGO.exists():
            logo_h = 0.3 * inch
            try:
                canvas.drawImage(
                    str(LOGO),
                    doc.leftMargin,
                    height - band_h + (band_h - logo_h) / 2,
                    height=logo_h,
                    width=logo_h,
                    mask="auto",
                    preserveAspectRatio=True,
                )
                text_x = doc.leftMargin + logo_h + 9
            except Exception:
                text_x = doc.leftMargin
        else:
            text_x = doc.leftMargin

        canvas.setFillColor(CHALK)
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawString(text_x, height - band_h / 2 - 1, self.doc_title)
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(colors.HexColor("#9AA5B1"))
        canvas.drawRightString(
            width - doc.rightMargin, height - band_h / 2 - 1, self.doc_subtitle
        )

        # Footer
        canvas.setStrokeColor(PLATE)
        canvas.setLineWidth(0.5)
        canvas.line(
            doc.leftMargin, 0.72 * inch, width - doc.rightMargin, 0.72 * inch
        )
        canvas.setFont("Helvetica", 6.8)
        canvas.setFillColor(FOG)
        canvas.drawString(doc.leftMargin, 0.55 * inch, DISCLAIMER)
        canvas.drawString(
            doc.leftMargin,
            0.4 * inch,
            f"Marapone Contracting · {EDITION} edition · built {BUILD_DATE}",
        )
        canvas.drawRightString(
            width - doc.rightMargin, 0.55 * inch, f"Page {canvas.getPageNumber()}"
        )
        canvas.restoreState()


_base = getSampleStyleSheet()

STYLES = {
    "h1": ParagraphStyle(
        "h1", parent=_base["Heading1"], fontName="Helvetica-Bold",
        fontSize=19, leading=23, spaceBefore=2, spaceAfter=9, textColor=INK,
    ),
    "h2": ParagraphStyle(
        "h2", parent=_base["Heading2"], fontName="Helvetica-Bold",
        fontSize=12, leading=15, spaceBefore=15, spaceAfter=6, textColor=INK,
    ),
    "h3": ParagraphStyle(
        "h3", parent=_base["Heading3"], fontName="Helvetica-Bold",
        fontSize=9.5, leading=12.5, spaceBefore=10, spaceAfter=3, textColor=HIVIZ,
    ),
    "body": ParagraphStyle(
        "body", parent=_base["BodyText"], fontName="Helvetica",
        fontSize=8.6, leading=12.2, spaceAfter=6, textColor=colors.HexColor("#22282E"),
        alignment=TA_LEFT,
    ),
    "small": ParagraphStyle(
        "small", parent=_base["BodyText"], fontName="Helvetica",
        fontSize=7.4, leading=10.2, spaceAfter=4, textColor=FOG,
    ),
    "cell": ParagraphStyle(
        "cell", fontName="Helvetica", fontSize=7.2, leading=9.2,
        textColor=colors.HexColor("#22282E"),
    ),
    "cellb": ParagraphStyle(
        "cellb", fontName="Helvetica-Bold", fontSize=7.2, leading=9.2, textColor=INK,
    ),
    "cellh": ParagraphStyle(
        "cellh", fontName="Helvetica-Bold", fontSize=7.2, leading=9.2, textColor=CHALK,
    ),
}


# Cell text comes straight out of the engine JSON and routinely contains "&"
# and quotes, so everything is escaped first. The authored copy in this file
# uses a handful of inline tags for emphasis, so those are restored afterwards
# — escape-then-restore rather than trusting the input, because a stray "<" in
# a source file must never be able to break a paragraph.
_INLINE_MARKUP = re.compile(r"&lt;(/?(?:b|i|br|font|sub|super)(?:\s[^&<>]*)?/?)&gt;")


def para(text, style="body"):
    escaped = (
        clean(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )
    return Paragraph(_INLINE_MARKUP.sub(r"<\1>", escaped), STYLES[style])


def title_block(title: str, standfirst: str, facts: list[tuple[str, str]]) -> list:
    """Title page furniture: what this file is, where it came from, when."""
    flow = [para(title, "h1"), para(standfirst, "body"), Spacer(1, 6)]
    rows = [[para(k, "cellb"), para(v, "cell")] for k, v in facts if v]
    table = Table(rows, colWidths=[1.65 * inch, None], hAlign="LEFT")
    table.setStyle(
        TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND", (0, 0), (-1, -1), BAND),
            ("BOX", (0, 0), (-1, -1), 0.5, PLATE),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.white),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ])
    )
    flow.append(table)
    return flow


def data_table(header: list[str], rows: list[list[str]], widths: list[float]) -> Table:
    body = [[para(h, "cellh") for h in header]]
    body += [[para(c, "cell") for c in row] for row in rows]
    table = Table(body, colWidths=widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, PLATE),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for i in range(1, len(body)):
        if i % 2 == 0:
            style.append(("BACKGROUND", (0, i), (-1, i), BAND))
    table.setStyle(TableStyle(style))
    return table


# ── Pack 1 — Building Code Compliance Library ────────────────────────────────

CODE_CSV_HEADER = [
    "check_id", "category", "requirement", "code_reference", "section",
    "plain_language_note", "last_checked", "source",
    "severity", "denial_risk", "correction",
]

# Which ruleset belongs to which folder. Province-level statute and municipal
# by-law are genuinely different jurisdictions and a buyer in Ottawa needs the
# first without the second, so they do not share a folder.
CODE_JURISDICTIONS = [
    {
        "folder": "canada/ontario",
        "name": "Ontario",
        "level": "Province",
        "files": ["obc_2025.json", "ontario_fire_code.json", "aoda_2025.json"],
    },
    {
        "folder": "canada/ontario/toronto",
        "name": "City of Toronto",
        "level": "Municipal",
        "files": ["toronto_zoning.json", "toronto_green_standard.json"],
    },
]


def rule_rows(ruleset: dict) -> list[list]:
    meta = ruleset.get("_meta", {})
    last = meta.get("last_updated", "")
    rows = []
    for rule in ruleset.get("rules", []):
        clause = clean(rule.get("clause"))
        rows.append([
            rule.get("id"),
            rule.get("category"),
            rule.get("description"),
            clause,
            section_of(clause),
            rule.get("plain_english"),
            last,
            rule.get("source") or meta.get("title"),
            rule.get("severity"),
            rule.get("denial_risk"),
            rule.get("correction"),
        ])
    return rows


def build_pack_code_library(engines: Path, out: Path) -> Path:
    rules_dir = engines / "Marapone-Takeoff-BP" / "resources" / "rules"
    root = out / "build" / "pack1-building-code-compliance-library"
    if root.exists():
        shutil.rmtree(root)
    root.mkdir(parents=True)

    index_rows: list[list] = []
    total = 0

    for juris in CODE_JURISDICTIONS:
        folder = root / juris["folder"]
        rows: list[list] = []
        sources: list[tuple[str, dict]] = []

        for filename in juris["files"]:
            ruleset = load(rules_dir / filename)
            rows += rule_rows(ruleset)
            meta = ruleset.get("_meta", {})
            sources.append((filename, meta))
            index_rows.append([
                juris["name"], juris["level"], meta.get("title"),
                meta.get("edition"), len(ruleset.get("rules", [])),
                meta.get("last_updated"), meta.get("scope"),
                f'{juris["folder"]}/checklist.csv',
            ])

        rows.sort(key=lambda r: (clean(r[1]), clean(r[0])))
        write_csv(folder / "checklist.csv", CODE_CSV_HEADER, rows)
        total += len(rows)
        _code_checklist_pdf(folder / "checklist.pdf", juris, rows, sources)

    write_csv(
        root / "index.csv",
        ["jurisdiction", "level", "ruleset", "edition", "checks",
         "last_checked", "scope", "file"],
        index_rows,
    )

    _code_start_here(root / "START-HERE.pdf", index_rows, total)
    write_text(root / "README.md", _code_readme(index_rows, total))
    return _zip(root, out / "marapone-building-code-compliance-library.zip")


def _code_checklist_pdf(path: Path, juris: dict, rows: list[list], sources) -> None:
    doc = PackDoc(
        path,
        f'Building Code Compliance — {juris["name"]}',
        f'{len(rows)} checks · {juris["level"]} level',
        wide=True,
    )
    facts = [
        ("Jurisdiction", f'{juris["name"]} ({juris["level"]} level)'),
        ("Checks in this file", str(len(rows))),
        ("Rulesets", "; ".join(clean(m.get("title")) for _, m in sources)),
        ("Editions", "; ".join(clean(m.get("edition")) for _, m in sources)),
        ("Last checked", "; ".join(clean(m.get("last_updated")) for _, m in sources)),
        ("Exported", BUILD_DATE),
        ("Machine-readable", "checklist.csv, same folder, same rows"),
    ]
    flow = title_block(
        f'Building Code Compliance Checklist — {juris["name"]}',
        "Every check below is a rule Blueprint Auditor applies to real drawings. "
        "The clause is the one the tool cites in its findings, so a check you "
        "read here is the check the software runs.",
        facts,
    )
    flow += [
        para("How to read a row", "h3"),
        para(
            "<b>Requirement</b> is the rule as the code states it. <b>Plain-language "
            "note</b> is what it means on a drawing. <b>Correction</b> is the fix "
            "that clears it. <b>Denial risk = yes</b> marks a check that has failed "
            "permit applications on its own.",
            "small",
        ),
        Spacer(1, 4),
        para(
            "Codes are amended. Every ruleset's last-checked date is on this page "
            "and in index.csv — verify any clause against the current consolidated "
            "text before relying on it in a submission.",
            "small",
        ),
        PageBreak(),
    ]

    by_category: dict[str, list[list]] = {}
    for row in rows:
        by_category.setdefault(clean(row[1]) or "General", []).append(row)

    # check_id gets room to sit on one or two clean lines: they are long, hyphenated
    # and unbreakable-looking, and an id split mid-word is not a usable reference.
    widths = [1.45 * inch, 1.45 * inch, 2.2 * inch, 0.72 * inch, 3.2 * inch, 0.52 * inch]
    for category, group in sorted(by_category.items()):
        flow.append(para(f"{category} — {len(group)} checks", "h2"))
        table_rows = [
            [r[0], r[3], r[2], r[8], r[5], "yes" if clean(r[9]) == "yes" else ""]
            for r in group
        ]
        flow.append(
            data_table(
                ["Check ID", "Code reference", "Requirement", "Severity",
                 "Plain-language note", "Denial"],
                table_rows,
                widths,
            )
        )
    doc.build(flow)


def _code_start_here(path: Path, index_rows: list[list], total: int) -> None:
    doc = PackDoc(path, "Building Code Compliance Library", "Start here", wide=False)
    flow = title_block(
        "Building Code Compliance Library",
        "The code-check ruleset behind Blueprint Auditor, exported as PDF and CSV. "
        f"{total} checks across {len(index_rows)} rulesets.",
        [
            ("Coverage", "Ontario (provincial) and City of Toronto (municipal)"),
            ("Checks", str(total)),
            ("Formats", "PDF to read, CSV to load"),
            ("Exported", BUILD_DATE),
            ("Licence", "One-time purchase. Use in your own projects, "
                        "spreadsheets and scripts. Do not resell as a data product."),
        ],
    )
    flow += [
        para("What is in the box", "h2"),
        para(
            "One folder per jurisdiction. Each holds <b>checklist.csv</b> (every "
            "check, one row each, UTF-8) and <b>checklist.pdf</b> (the same rows "
            "grouped by category and laid out to read). <b>index.csv</b> at the "
            "root lists every ruleset with its edition and last-checked date.",
        ),
        para("Coverage, stated plainly", "h2"),
        para(
            "This edition covers <b>Ontario</b> and the <b>City of Toronto</b> "
            "municipal layer. Those are the jurisdictions the Blueprint Auditor "
            "engine carries rule data for, and this pack is an export of that "
            "engine rather than a separate research product — so it covers "
            "exactly what the tool covers, no more.",
        ),
        para(
            "If you build outside Ontario, the structure is still useful as a "
            "checklist spine (the categories — egress, fire separation, "
            "accessibility, zoning envelope — are consistent across codes), but "
            "the clause numbers and dimensions are Ontario's. Do not apply them "
            "elsewhere.",
        ),
        para("The CSV schema", "h2"),
        para(
            "Identical in every jurisdiction folder, so files concatenate cleanly:",
            "small",
        ),
        data_table(
            ["Column", "What it holds"],
            [
                ["check_id", "Stable identifier. Safe to key on."],
                ["category", "Egress, fire separation, accessibility, setbacks, and so on."],
                ["requirement", "The rule as the code states it."],
                ["code_reference", "The clause cited, exactly as Blueprint Auditor cites it."],
                ["section", "Numeric section parsed out of the clause, for sorting."],
                ["plain_language_note", "What the rule means on a drawing."],
                ["last_checked", "When this ruleset was last reviewed against source."],
                ["source", "The code or by-law the check comes from."],
                ["severity", "CRITICAL / MAJOR / MINOR as the engine ranks it."],
                ["denial_risk", "yes where the check has failed permit applications alone."],
                ["correction", "The fix that clears the check."],
            ],
            [1.5 * inch, None],
        ),
        para("Rulesets in this edition", "h2"),
        data_table(
            ["Jurisdiction", "Ruleset", "Edition", "Checks", "Last checked"],
            [[r[0], r[2], r[3], r[4], r[5]] for r in index_rows],
            [1.1 * inch, 2.5 * inch, 1.6 * inch, 0.6 * inch, 0.85 * inch],
        ),
        para("Updates", "h2"),
        para(
            "Codes are amended on their own schedule, not ours. This is a dated "
            "export, not a subscription: when a ruleset changes materially the next "
            "edition ships as a new dated file rather than a silent overwrite, so a "
            "bid you defended against this edition stays defensible.",
        ),
    ]
    doc.build(flow)


def _code_readme(index_rows: list[list], total: int) -> str:
    table = "\n".join(
        f"| {r[0]} | {r[2]} | {r[3]} | {r[4]} | {r[5]} |" for r in index_rows
    )
    return f"""# Building Code Compliance Library

The code-check ruleset behind [Blueprint Auditor](https://maraponecontracting.com/construction/blueprint-auditor),
exported as PDF and CSV. **{total} checks** across {len(index_rows)} rulesets.

Exported {BUILD_DATE}.

## Coverage

**Ontario (provincial) and the City of Toronto (municipal layer).**

That is what the Blueprint Auditor engine carries rule data for. This pack is a
straight export of that engine, so it covers exactly what the tool covers.

| Jurisdiction | Ruleset | Edition | Checks | Last checked |
|---|---|---|---|---|
{table}

If you build outside Ontario: the category spine (egress, fire separation,
accessibility, zoning envelope) transfers as a checklist structure, but the
clause numbers and dimensions do not. Do not apply them to another code.

## Layout

```
index.csv                              every ruleset, edition, last-checked date
START-HERE.pdf                         how to use this, schema, licence
canada/ontario/checklist.csv           Ontario Building Code, Fire Code, AODA
canada/ontario/checklist.pdf
canada/ontario/toronto/checklist.csv   Zoning By-law 569-2013, Green Standard
canada/ontario/toronto/checklist.pdf
```

## CSV schema

`check_id, category, requirement, code_reference, section, plain_language_note,
last_checked, source, severity, denial_risk, correction`

UTF-8 with BOM, CRLF line endings, identical in every jurisdiction folder — the
files concatenate without reconciling headers.

The first eight columns are the documented schema. `severity`, `denial_risk` and
`correction` are appended because the engine carries them and they are the most
useful columns in practice.

## Disclaimer

Reference material only — not a substitute for review by a licensed
professional in your jurisdiction. Codes are amended; check the `last_checked`
date on any row before relying on it in a submission.
"""


# ── Pack 2 — Assembly Property Schema ────────────────────────────────────────


def build_pack_assembly_schema(engines: Path, out: Path) -> Path:
    knowledge = engines / "SpecChecker" / "resources" / "knowledge"
    attrs = load(knowledge / "attribute_rules.json")
    divisions = load(knowledge / "csi_divisions.json")

    root = out / "build" / "pack2-assembly-property-schema"
    if root.exists():
        shutil.rmtree(root)
    root.mkdir(parents=True)
    na = root / "north-america"

    properties = attrs["attributes"]
    assemblies = attrs["assemblies"]
    meta = attrs.get("_meta", {})

    # Which properties apply to each assembly — the join the engine does at
    # runtime, materialised so a buyer does not have to redo it.
    by_assembly: dict[str, list[str]] = {a["key"]: [] for a in assemblies}
    for prop in properties:
        for key in prop.get("assemblies", []):
            by_assembly.setdefault(key, []).append(prop["key"])

    section_for: dict[str, list[str]] = {}
    for sec in divisions.get("sections", []):
        for key in sec.get("assemblies", []):
            section_for.setdefault(key, []).append(
                f'{sec["number"]} {sec["title"]}'
            )

    div_name = {d["code"]: d["name"] for d in divisions.get("divisions", [])}

    # A machine-readable file that keeps the join rather than flattening it away.
    export = {
        "_meta": {
            "pack": "Marapone Assembly Property Schema",
            "standard": divisions.get("_meta", {}).get("standard"),
            "source_engine": "Marapone SpecChecker",
            "source_revision": meta.get("revision"),
            "source_region": meta.get("region"),
            "exported": BUILD_DATE,
            "properties": len(properties),
            "assemblies": len(assemblies),
            "divisions": len(divisions.get("divisions", [])),
            "sections": len(divisions.get("sections", [])),
            "disclaimer": DISCLAIMER,
        },
        "properties": properties,
        "assemblies": [
            {
                **assembly,
                "division_name": div_name.get(assembly.get("division"), ""),
                "properties": by_assembly.get(assembly["key"], []),
                "csi_sections": section_for.get(assembly["key"], []),
            }
            for assembly in assemblies
        ],
        "csi_divisions": divisions.get("divisions", []),
        "csi_sections": divisions.get("sections", []),
    }
    na.mkdir(parents=True, exist_ok=True)
    (na / "csi-masterformat-assemblies.json").write_text(
        json.dumps(export, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )

    write_csv(
        na / "csi-masterformat-assemblies.csv",
        ["assembly_key", "assembly_label", "csi_division", "csi_division_name",
         "csi_sections", "property_count", "properties", "keywords",
         "drawing_tag_prefixes", "schedules", "source", "last_checked"],
        [[
            a["key"], a.get("label"), a.get("division"),
            div_name.get(a.get("division"), ""),
            section_for.get(a["key"], []),
            len(by_assembly.get(a["key"], [])),
            by_assembly.get(a["key"], []),
            a.get("keywords"), a.get("tag_prefixes"), a.get("schedules"),
            "Marapone SpecChecker attribute_rules.json", meta.get("revision"),
        ] for a in assemblies],
    )

    write_csv(
        na / "assembly-properties.csv",
        ["property_key", "label", "kind", "unit", "direction", "tolerance",
         "critical", "applies_to_assemblies", "accepted_values",
         "why_it_matters", "rfi_note", "source", "last_checked"],
        [[
            p["key"], p.get("label"), p.get("kind"), p.get("unit_label") or p.get("unit"),
            {"min": "specified value is a minimum",
             "max": "specified value is a maximum",
             "exact": "must match within tolerance"}.get(p.get("direction"), ""),
            p.get("tolerance", p.get("tolerance_pct", "")),
            p.get("critical"),
            p.get("assemblies"),
            [v.get("value") for v in p.get("values", [])],
            p.get("why"), p.get("rfi_note"),
            "Marapone SpecChecker attribute_rules.json", meta.get("revision"),
        ] for p in properties],
    )

    write_csv(
        na / "csi-divisions.csv",
        ["division_code", "division_name", "short_name", "audited"],
        [[d["code"], d["name"], d.get("short"), d.get("audited")]
         for d in divisions.get("divisions", [])],
    )
    write_csv(
        na / "csi-sections.csv",
        ["section_number", "section_title", "assemblies"],
        [[s["number"], s["title"], s.get("assemblies")]
         for s in divisions.get("sections", [])],
    )

    _schema_reference_pdf(root / "schema-reference.pdf", properties, assemblies,
                          divisions, by_assembly, meta)
    write_text(root / "README.md",
               _schema_readme(properties, assemblies, divisions, meta))
    return _zip(root, out / "marapone-assembly-property-schema.zip")


def _schema_reference_pdf(path, properties, assemblies, divisions, by_assembly, meta):
    doc = PackDoc(path, "Assembly Property Schema",
                  f"{len(properties)} properties · {len(assemblies)} assemblies")
    flow = title_block(
        "Assembly Property Schema",
        "The structured assembly taxonomy behind SpecChecker: the properties the "
        "engine compares between a specification and a drawing, and the assemblies "
        "each one applies to.",
        [
            ("Standard", clean(divisions.get("_meta", {}).get("standard"))),
            ("Properties", str(len(properties))),
            ("Assemblies", str(len(assemblies))),
            ("CSI divisions / sections",
             f'{len(divisions.get("divisions", []))} / {len(divisions.get("sections", []))}'),
            ("Source revision", clean(meta.get("revision"))),
            ("Exported", BUILD_DATE),
        ],
    )
    flow += [
        para("Why only these properties", "h2"),
        para(clean(meta.get("note")) or
             "A spec clause and a drawing callout only ever conflict about a "
             "measurable property of a named assembly."),
        para(
            "<b>Numeric</b> properties carry a unit, a direction (is the specified "
            "figure a floor or a ceiling?) and a tolerance. <b>Designation</b> "
            "properties carry a ranked value list, so the engine can tell a "
            "substitution up from a substitution down. <b>Critical</b> marks the "
            "properties that cost real money when they are wrong — the ones that "
            "are fixed at fabrication, batching or inspection and cannot be "
            "corrected on site.",
        ),
        PageBreak(),
        para("The properties", "h2"),
    ]

    for prop in properties:
        unit = clean(prop.get("unit_label") or prop.get("unit"))
        kind = clean(prop.get("kind"))
        bits = [kind]
        if unit:
            bits.append(f"measured in {unit}")
        direction = {"min": "specified value is a minimum",
                     "max": "specified value is a maximum",
                     "exact": "must match within tolerance"}.get(prop.get("direction"))
        if direction:
            bits.append(direction)
        if prop.get("critical"):
            bits.append("critical")

        block = [
            para(f'{prop.get("label")} — <font size="7">{prop["key"]}</font>', "h3"),
            para(" · ".join(b for b in bits if b), "small"),
            para(prop.get("why")),
        ]
        if prop.get("values"):
            ranked = sorted(prop["values"], key=lambda v: v.get("rank", 0))
            block.append(para(
                "<b>Accepted values, lowest to highest:</b> "
                + ", ".join(clean(v.get("value")) for v in ranked), "small"))
        block.append(para(
            f'<b>Applies to:</b> {clean(prop.get("assemblies"))}', "small"))
        block.append(para(f'<b>Standard RFI wording:</b> {clean(prop.get("rfi_note"))}',
                          "small"))
        flow.append(KeepTogether(block))

    flow += [PageBreak(), para("Assemblies and the properties checked on each", "h2")]
    div_name = {d["code"]: d["short"] or d["name"] for d in divisions.get("divisions", [])}
    flow.append(data_table(
        ["Assembly", "Div", "Division", "Properties checked"],
        [[a.get("label"), a.get("division"), div_name.get(a.get("division"), ""),
          ", ".join(by_assembly.get(a["key"], [])) or "—"]
         for a in assemblies],
        [1.25 * inch, 0.35 * inch, 1.35 * inch, None],
    ))
    doc.build(flow)


def _schema_readme(properties, assemblies, divisions, meta) -> str:
    return f"""# Assembly Property Schema

The structured assembly taxonomy behind [SpecChecker](https://maraponecontracting.com/construction/specchecker):
**{len(properties)} properties** across **{len(assemblies)} assemblies**, mapped to
{len(divisions.get('divisions', []))} CSI MasterFormat divisions and
{len(divisions.get('sections', []))} sections.

Exported {BUILD_DATE} from SpecChecker revision {clean(meta.get('revision'))}.

## Standard covered

**{clean(divisions.get('_meta', {}).get('standard'))} — North America.**

This edition ships the CSI MasterFormat-aligned tree only. The SpecChecker
engine is built on CSI/CSC MasterFormat and does not carry an EN/Eurocode
classification, so there is no Eurocode file in this pack. Shipping an empty
`/europe/` folder, or a machine-translated CSI tree relabelled as EN, would be
worse than shipping neither — the two are genuinely different classification
systems and a mapping between them is a research project, not an export.

The *properties* themselves are largely standard-neutral: a fire-resistance
rating, a compressive strength and a U-value are the same measurable quantities
under any code. What is North American is the CSI division/section spine and
some of the accepted designation values (CSA steel grades, ANSI hardware
grades). Both are labelled at the row level, so filtering them is one column
away.

## Layout

```
north-america/csi-masterformat-assemblies.json   full schema, joins intact
north-america/csi-masterformat-assemblies.csv    one row per assembly
north-america/assembly-properties.csv            one row per property
north-america/csi-divisions.csv                  {len(divisions.get('divisions', []))} divisions
north-america/csi-sections.csv                   {len(divisions.get('sections', []))} sections
schema-reference.pdf                             every property in plain language
```

## The join

Properties and assemblies are many-to-many: `fire_rating` applies to nine
assemblies, and a `door` carries five properties. The JSON keeps that join in
both directions (`properties[].assemblies` and `assemblies[].properties`); the
CSVs materialise it so a spreadsheet does not have to.

## Reading a property

- **kind** — `numeric` (has a unit and a tolerance) or `designation` (has a
  ranked list of accepted values).
- **direction** — whether the specified figure is a floor, a ceiling, or a
  target to match. Getting this backwards inverts the check.
- **critical** — fixed at fabrication, batching or inspection; cannot be
  corrected on site once wrong.
- **accepted_values** — ranked low to high, so a substitution can be scored as
  an upgrade or a downgrade rather than just "different".

## Disclaimer

Reference material only — not a substitute for review by a licensed
professional in your jurisdiction.
"""


# ── Pack 3 — Estimating Rate Pack, 2026 Edition ──────────────────────────────


def build_pack_rate(engines: Path, out: Path) -> Path:
    pricing = engines / "AI-Estimator" / "resources" / "pricing"
    juris = load(pricing / "jurisdiction_on_toronto.json")
    soft = load(pricing / "toronto_soft_costs.json")
    factors = load(pricing / "location_factors.json")
    risk = load(engines / "AI-Estimator" / "resources" / "risk" / "risk_rules.json")

    root = out / "build" / "pack3-estimating-rate-pack-2026"
    if root.exists():
        shutil.rmtree(root)
    root.mkdir(parents=True)

    jmeta = juris.get("_meta", {})
    smeta = soft.get("_meta", {})
    folder = root / "canada" / "ontario" / "toronto"

    tax = juris.get("tax", {})
    burden = juris.get("labour_burden", {})
    act = juris.get("construction_act", {})
    bond = juris.get("bonding_insurance", {})

    # index.csv — the flat "what rate applies where" table.
    index_rows = [
        ["Ontario", "Province", "HST (combined)", pct(tax.get("hst_pct")),
         jmeta.get("revision"), clean(tax.get("citation"))],
        ["Ontario", "Province", "HST — federal portion",
         pct(tax.get("hst_federal_portion")), jmeta.get("revision"),
         clean(tax.get("citation"))],
        ["Ontario", "Province", "HST — provincial portion",
         pct(tax.get("hst_provincial_portion")), jmeta.get("revision"),
         clean(tax.get("citation"))],
        ["Ontario", "Province", "Employer Health Tax", pct(burden.get("eht_pct")),
         jmeta.get("revision"), clean(burden.get("eht_citation"))],
        ["Ontario", "Province", "CPP — employer", pct(burden.get("cpp_employer_pct")),
         jmeta.get("revision"), "Canada Pension Plan — employer contribution"],
        ["Ontario", "Province", "EI — employer", pct(burden.get("ei_employer_pct")),
         jmeta.get("revision"), "Employment Insurance — employer contribution"],
        ["Ontario", "Province", "Vacation pay (minimum)",
         pct(burden.get("vacation_pay_pct")), jmeta.get("revision"),
         "Employment Standards Act (Ontario)"],
        ["Ontario", "Province", "Statutory holdback", pct(act.get("statutory_holdback_pct")),
         jmeta.get("revision"), clean(act.get("citation"))],
        ["Ontario", "Province", "Bond premium (% of contract)",
         pct(bond.get("bond_premium_pct_of_contract")), jmeta.get("revision"),
         "Surety bond premium — performance and L&M"],
    ]
    for group, rate in (burden.get("wsib_rate_group") or {}).items():
        if isinstance(rate, (int, float)):
            index_rows.append([
                "Ontario", "Province", f"WSIB premium — {humanise(group)}",
                f"${rate:g} per $100 insurable earnings", jmeta.get("revision"),
                clean(burden["wsib_rate_group"].get("citation")),
            ])
    index_rows.append([
        "City of Toronto", "Municipal", "Building permit fees",
        "see canada/ontario/toronto/rates.csv", smeta.get("rate_year"),
        "City of Toronto Municipal Code Ch. 363",
    ])
    index_rows.append([
        "City of Toronto", "Municipal", "Development charges",
        "see canada/ontario/toronto/rates.csv", smeta.get("rate_year"),
        "City of Toronto Development Charges By-law",
    ])

    write_csv(
        root / "index.csv",
        ["jurisdiction", "level", "tax_type", "rate", "effective_date", "source"],
        index_rows,
    )

    # The full jurisdiction rate table, flattened one leaf per row.
    rate_rows: list[list] = []

    def flatten(node, path: list[str], source: str, effective: str):
        if isinstance(node, dict):
            local_source = clean(node.get("citation")) or source
            local_eff = clean(node.get("rate_year") or node.get("effective")) or effective
            for key, value in node.items():
                if key in ("citation", "_note"):
                    continue
                flatten(value, path + [key], local_source, local_eff)
        elif isinstance(node, list):
            rate_rows.append([" / ".join(path), clean(node), "", source, effective])
        else:
            label = path[-1]
            value = node
            if isinstance(value, float) and label.endswith("_pct"):
                value = pct(value)
            rate_rows.append([" / ".join(path[:-1]), label, clean(value), source, effective])

    for section, node in juris.items():
        if section == "_meta":
            continue
        flatten(node, [section], clean(jmeta.get("revision")), clean(jmeta.get("revision")))
    for section, node in soft.items():
        if section == "_meta":
            continue
        flatten(node, [section], "; ".join(smeta.get("sources", [])),
                clean(smeta.get("rate_year")))

    write_csv(
        folder / "rates.csv",
        ["group", "item", "value", "source", "effective"],
        rate_rows,
    )

    cities = factors.get("cities", {})
    write_csv(
        folder / "city-cost-index.csv",
        ["city", "region", "material_factor", "labour_factor", "equipment_factor",
         "subcontract_factor", "note", "baseline", "source"],
        [[c.get("name"), c.get("region"), c.get("material"), c.get("labor"),
          c.get("equipment"), c.get("subcontract"), c.get("note"),
          factors.get("_meta", {}).get("base_city"),
          f'Marapone AI Estimator location_factors.json rev {clean(factors.get("_meta", {}).get("revision"))}']
         for c in cities.values()],
    )

    rules = risk.get("rules", [])
    write_csv(
        root / "contingency-risk-framework.csv",
        ["rule_id", "name", "category", "severity", "contingency_pct", "applied_to",
         "trigger", "explanation", "mitigation", "confidence_penalty"],
        [[r.get("id"), r.get("name"), r.get("category"), r.get("severity"),
          pct(r.get("contingency_pct")), r.get("basis"), r.get("trigger"),
          r.get("explanation"), r.get("mitigation"), r.get("confidence_penalty")]
         for r in rules],
    )

    _rates_pdf(folder / "rates.pdf", juris, soft, cities, factors)
    _contingency_pdf(root / "contingency-risk-framework.pdf", risk)
    _rate_start_here(root / "START-HERE.pdf", index_rows, rules, cities)
    write_text(root / "README.md",
               _rate_readme(juris, soft, rules, cities, rate_rows))
    return _zip(root, out / "marapone-estimating-rate-pack-2026.zip")


MUNICIPAL_NOTE = (
    "<b>Precision note.</b> Permits and development charges are set municipally. "
    "The figures in this file are the <b>City of Toronto</b> schedule, not an "
    "Ontario average — they do not apply in Mississauga, Hamilton or Ottawa, "
    "each of which sets its own. For a project outside Toronto, use the "
    "provincial rows (HST, WSIB, holdback, prompt payment) as-is and obtain the "
    "permit and development-charge schedule from that local authority."
)


def _rates_pdf(path, juris, soft, cities, factors):
    jmeta = juris.get("_meta", {})
    smeta = soft.get("_meta", {})
    doc = PackDoc(path, "Estimating Rates — Toronto, Ontario",
                  f'{EDITION} edition · {clean(jmeta.get("currency"))}')
    flow = title_block(
        "Estimating Rate Pack — City of Toronto, Ontario",
        clean(jmeta.get("description")),
        [
            ("Jurisdiction", clean(jmeta.get("jurisdiction"))),
            ("Currency", clean(jmeta.get("currency"))),
            ("Source revision", clean(jmeta.get("revision"))),
            ("Municipal rate year", clean(smeta.get("rate_year"))),
            ("Data embedded", clean(smeta.get("embedded_at_build"))),
            ("Exported", BUILD_DATE),
            ("Machine-readable", "rates.csv, same folder, every row"),
        ],
    )
    flow += [para(MUNICIPAL_NOTE, "small"), Spacer(1, 3),
             para(clean(smeta.get("update_policy")), "small")]

    tax = juris.get("tax", {})
    burden = juris.get("labour_burden", {})
    act = juris.get("construction_act", {})
    bond = juris.get("bonding_insurance", {})

    flow += [
        para("Tax", "h2"),
        data_table(
            ["Item", "Rate", "Source"],
            [
                ["HST — combined", pct(tax.get("hst_pct")), clean(tax.get("citation"))],
                ["HST — federal portion", pct(tax.get("hst_federal_portion")), ""],
                ["HST — provincial portion", pct(tax.get("hst_provincial_portion")), ""],
            ],
            [2.0 * inch, 1.2 * inch, None],
        ),
        para(clean(tax.get("note")), "small"),
    ]
    rebate = tax.get("new_housing_rebate") or {}
    if rebate:
        flow.append(para(
            f'<b>New housing rebate.</b> {clean(rebate.get("applies_to"))}. '
            f'{clean(rebate.get("note"))}', "small"))

    flow += [
        para("Labour burden", "h2"),
        para(clean(burden.get("_note")), "small"),
        data_table(
            ["Item", "Rate", "Source"],
            [["CPP — employer", pct(burden.get("cpp_employer_pct")), ""],
             ["EI — employer", pct(burden.get("ei_employer_pct")), ""],
             ["Employer Health Tax", pct(burden.get("eht_pct")),
              clean(burden.get("eht_citation"))],
             ["Vacation pay (minimum)", pct(burden.get("vacation_pay_pct")),
              clean(burden.get("note"))],
             ["Statutory holidays", clean(burden.get("statutory_holidays")), ""]],
            [2.0 * inch, 1.2 * inch, None],
        ),
    ]
    wsib = burden.get("wsib_rate_group") or {}
    flow += [
        para("WSIB premium by rate group", "h3"),
        data_table(
            ["Rate group", f'Premium ({clean(wsib.get("unit"))})'],
            [[humanise(k), f"${v:g}"]
             for k, v in wsib.items() if isinstance(v, (int, float))],
            [2.6 * inch, None],
        ),
        para(clean(wsib.get("citation")), "small"),
    ]

    flow += [
        para("Construction Act — what it does to cash flow", "h2"),
        para(clean(act.get("citation")), "small"),
        data_table(
            ["Item", "Value"],
            [["Statutory holdback", pct(act.get("statutory_holdback_pct"))],
             ["Holdback release", f'{clean(act.get("holdback_release_days"))} days'],
             ["Prompt payment — owner to contractor",
              f'{clean((act.get("prompt_payment") or {}).get("owner_to_contractor_days"))} days'],
             ["Prompt payment — contractor to sub",
              f'{clean((act.get("prompt_payment") or {}).get("contractor_to_sub_days"))} days'],
             ["Lien preservation", f'{clean(act.get("lien_preservation_days"))} days'],
             ["Lien perfection", f'{clean(act.get("lien_perfection_days"))} days']],
            [2.8 * inch, None],
        ),
        para(clean(act.get("cost_impact")), "small"),
        para(clean((act.get("prompt_payment") or {}).get("note")), "small"),
    ]

    flow += [
        para("Bonding and insurance", "h2"),
        data_table(
            ["Item", "Value"],
            [[humanise(k),
              pct(v) if isinstance(v, float) and "pct" in k else clean(v)]
             for k, v in bond.items() if not isinstance(v, (dict, list))],
            [2.8 * inch, None],
        ),
    ]

    flow.append(PageBreak())
    flow.append(para("Municipal soft costs — City of Toronto", "h2"))
    flow.append(para(clean(smeta.get("description")), "small"))
    flow.append(para(MUNICIPAL_NOTE, "small"))
    for key, node in soft.items():
        if key == "_meta" or not isinstance(node, dict):
            continue
        rows = []
        for sub, value in node.items():
            if isinstance(value, (dict, list)):
                rows.append([humanise(sub), clean(value)])
            else:
                rows.append([humanise(sub),
                             pct(value) if isinstance(value, float) and "pct" in sub
                             else clean(value)])
        if rows:
            flow.append(para(humanise(key), "h3"))
            flow.append(data_table(["Item", "Value"], rows, [2.2 * inch, None]))
    flow.append(para("<b>Sources.</b> " + "; ".join(smeta.get("sources", [])), "small"))

    flow += [
        PageBreak(),
        para("City cost index — Toronto/GTA", "h2"),
        para(clean(factors.get("_meta", {}).get("description")), "small"),
        data_table(
            ["City", "Material", "Labour", "Equipment", "Subcontract", "Note"],
            [[c.get("name"), c.get("material"), c.get("labor"), c.get("equipment"),
              c.get("subcontract"), c.get("note")] for c in cities.values()],
            [1.05 * inch, 0.62 * inch, 0.58 * inch, 0.7 * inch, 0.75 * inch, None],
        ),
        para(
            "These factors are GTA-relative and all sit inside Ontario. They "
            "adjust a Toronto rate to a nearby municipality; they are not "
            "national or international cost indices and will not carry a rate to "
            "another province or country.", "small",
        ),
    ]
    doc.build(flow)


def _contingency_pdf(path, risk):
    rules = risk.get("rules", [])
    bands = risk.get("contingency_bands", {})
    doc = PackDoc(path, "Contingency & Risk Framework",
                  f"{len(rules)} rules · jurisdiction-neutral")
    flow = title_block(
        "Contingency and Risk Framework",
        "The rules the AI Estimator evaluates against an estimate to produce a "
        "recommended contingency it can defend in a bid review. Each rule that "
        "fires contributes a named percentage to the total, applied to a stated "
        "basis — not one opinion-shaped number.",
        [
            ("Rules", str(len(rules))),
            ("Jurisdiction", "None — this framework is structural and travels"),
            ("Source", "Marapone AI Estimator risk_rules.json"),
            ("Exported", BUILD_DATE),
            ("Machine-readable", "contingency-risk-framework.csv"),
        ],
    )
    flow += [
        para("How the total is assembled", "h3"),
        para(clean(risk.get("_meta", {}).get("description")), "small"),
    ]
    if bands:
        flow += [
            para("Contingency bands", "h2"),
            data_table(
                ["Band", "Definition"],
                [[humanise(k), clean(v)] for k, v in bands.items()],
                [1.5 * inch, None],
            ),
        ]

    by_category: dict[str, list[dict]] = {}
    for rule in rules:
        by_category.setdefault(clean(rule.get("category")) or "general", []).append(rule)

    for category, group in sorted(by_category.items()):
        flow.append(para(f"{humanise(category)} — {len(group)} rules", "h2"))
        for rule in group:
            flow.append(KeepTogether([
                para(f'{rule.get("name")} — <font size="7">{rule.get("id")}</font>', "h3"),
                para(
                    f'Severity <b>{clean(rule.get("severity"))}</b> · adds '
                    f'<b>{pct(rule.get("contingency_pct"))}</b> to contingency, '
                    f'applied to <b>{clean(rule.get("basis")).replace("_", " ")}</b>'
                    + (f' · confidence penalty {rule.get("confidence_penalty")}'
                       if rule.get("confidence_penalty") else ""),
                    "small",
                ),
                para(rule.get("explanation")),
                para(f'<b>Mitigation.</b> {clean(rule.get("mitigation"))}', "small"),
            ]))
    doc.build(flow)


def _rate_start_here(path, index_rows, rules, cities):
    doc = PackDoc(path, "Estimating Rate Pack", f"{EDITION} edition · start here")
    flow = title_block(
        f"Estimating Rate Pack — {EDITION} Edition",
        "Statutory and municipal costing inputs behind the AI Estimator, plus the "
        "contingency framework it applies to every estimate.",
        [
            ("Coverage", "Ontario (statutory) and City of Toronto (municipal)"),
            ("Currency", "CAD"),
            ("Contingency rules", str(len(rules))),
            ("City cost index", f"{len(cities)} Toronto/GTA municipalities"),
            ("Edition", f"{EDITION} — re-released annually, not silently patched"),
            ("Exported", BUILD_DATE),
        ],
    )
    flow += [
        para("What is in the box", "h2"),
        data_table(
            ["File", "What it holds"],
            [
                ["index.csv", "Every headline rate with its source and effective date."],
                ["canada/ontario/toronto/rates.pdf", "The full rate table, laid out to read."],
                ["canada/ontario/toronto/rates.csv", "The same figures, one leaf per row."],
                ["canada/ontario/toronto/city-cost-index.csv",
                 f"{len(cities)} GTA municipalities as factors against Toronto = 1.000."],
                ["contingency-risk-framework.pdf", f"All {len(rules)} rules, written out."],
                ["contingency-risk-framework.csv", "The same rules, machine-readable."],
            ],
            [2.5 * inch, None],
        ),
        para("Coverage, stated plainly", "h2"),
        para(
            "The statutory rows — HST, CPP, EI, Employer Health Tax, WSIB, "
            "Construction Act holdback and prompt-payment clocks — apply across "
            "<b>Ontario</b>. The permit, development-charge, parkland and community "
            "benefits figures are the <b>City of Toronto</b> schedule specifically.",
        ),
        para(MUNICIPAL_NOTE),
        para(
            "This edition does not carry other provinces, US states or EU VAT. "
            "Those are not thin entries to be topped up — the AI Estimator engine "
            "holds one jurisdiction file, Toronto/Ontario, and this pack is an "
            "export of it. The contingency framework is the exception: it is "
            "structural, carries no jurisdiction-bound numbers, and is written to "
            "be applied to an estimate anywhere.",
        ),
        para("Why this edition is dated", "h2"),
        para(
            "This is the one pack that goes stale on a clock. Tax rates change at "
            "budgets, WSIB premiums are set annually, and Toronto's development "
            f"charge schedule is revised every year. The {EDITION} edition is a "
            "snapshot with its dates attached. Next year's rates ship as a "
            "<b>2027 Edition</b> — a new file you choose to buy, not a silent "
            "overwrite of the numbers you priced a live bid against.",
        ),
        para("Verify before you bid", "h2"),
        para(
            "Every figure carries its source and its effective date. That is there "
            "to be used: on a submitted bid, check the rate against the current "
            "published schedule. This pack is a costing input and a starting "
            "position, not legal or tax advice.",
        ),
    ]
    doc.build(flow)


def _rate_readme(juris, soft, rules, cities, rate_rows) -> str:
    jmeta = juris.get("_meta", {})
    smeta = soft.get("_meta", {})
    sources = "\n".join(f"- {s}" for s in smeta.get("sources", []))
    return f"""# Estimating Rate Pack — {EDITION} Edition

Statutory and municipal costing inputs behind the
[AI Estimator](https://maraponecontracting.com/construction/ai-estimator), plus
the **{len(rules)}-rule contingency framework** it applies to every estimate.

{len(rate_rows)} rate rows · {len(cities)} GTA municipalities · exported {BUILD_DATE}.

## Coverage

**Ontario (statutory) and the City of Toronto (municipal).** Currency: CAD.

| Layer | Applies to | Examples |
|---|---|---|
| Statutory | All of Ontario | HST, CPP, EI, Employer Health Tax, WSIB rate groups, Construction Act holdback and prompt-payment clocks |
| Municipal | City of Toronto only | Building permit fees, development charges, parkland dedication, community benefits charge |
| Structural | Anywhere | The {len(rules)}-rule contingency and risk framework — no jurisdiction-bound numbers |

**Precision note.** Permits and development charges are set municipally. The
figures here are the City of Toronto schedule, *not* an Ontario average — they
do not apply in Mississauga, Hamilton or Ottawa. For a project outside Toronto,
use the provincial rows as-is and obtain the permit and development-charge
schedule from that local authority.

This edition does not carry other provinces, US state sales tax or EU VAT. The
AI Estimator engine holds one jurisdiction file — Toronto/Ontario — and this
pack is an export of it.

## Layout

```
index.csv                                   headline rates, source, effective date
START-HERE.pdf                              how to use this, coverage, licence
contingency-risk-framework.pdf              all {len(rules)} rules, written out
contingency-risk-framework.csv              the same rules, machine-readable
canada/ontario/toronto/rates.pdf            full rate table, laid out to read
canada/ontario/toronto/rates.csv            {len(rate_rows)} rows, one leaf each
canada/ontario/toronto/city-cost-index.csv  {len(cities)} GTA municipalities
```

## Versioning

This is the one pack that goes stale on a clock. Next year's figures ship as a
**2027 Edition** — a new dated file, not a silent overwrite. A bid you defended
against the {EDITION} edition stays defensible.

Source revision {clean(jmeta.get('revision'))}; municipal rate year
{clean(smeta.get('rate_year'))}; data embedded {clean(smeta.get('embedded_at_build'))}.

## Sources

{sources}

## Disclaimer

Reference material only — not a substitute for review by a licensed
professional in your jurisdiction. Costing input, not legal or tax advice.
Verify every figure against the current published schedule before relying on it
in a submitted bid.
"""


# ── Pack 4 — Bid-Levelling Template ──────────────────────────────────────────

BIDDER_COLS = 3


def build_pack_bid_leveling(engines: Path, out: Path) -> Path:
    data = load(engines / "BidLeveler" / "resources" / "knowledge" / "trade_scopes.json")
    juris = load(engines / "AI-Estimator" / "resources" / "pricing" /
                 "jurisdiction_on_toronto.json")

    root = out / "build" / "pack4-bid-leveling-template"
    if root.exists():
        shutil.rmtree(root)
    root.mkdir(parents=True)

    meta = data.get("_meta", {})
    universal = data.get("universal", [])
    trades = data.get("trades", [])

    # One flat list: universal elements apply to every trade, trade elements to
    # one. 18 + 137 = 155 lines a sub can quietly leave out of a quote.
    rows: list[list] = []
    for element in universal:
        rows.append([
            element.get("key"), "ALL", "Applies to every trade", element.get("csi"),
            element.get("label"), "universal", element.get("critical"),
            pct(element.get("percent_of_trade")), element.get("description"),
            element.get("terms"),
        ])
    for trade in trades:
        for element in trade.get("elements", []):
            rows.append([
                element.get("key"), trade.get("code"), trade.get("name"),
                element.get("csi"), element.get("label"), "trade",
                element.get("critical"), pct(element.get("percent_of_trade")),
                element.get("description"), element.get("terms"),
            ])

    header = ["element_key", "trade_code", "trade_package", "csi_code", "scope_element",
              "scope_type", "critical", "typical_percent_of_trade", "what_to_look_for",
              "quote_terms"]
    write_csv(root / "scope-checklist.csv", header, rows)

    _bid_leveling_xlsx(root / "marapone-bid-leveling-template.xlsx",
                       rows, header, trades, universal, juris, meta)
    _bid_how_to_pdf(root / "how-to-use.pdf", rows, trades, universal, juris, meta)
    write_text(root / "README.md", _bid_readme(rows, trades, universal, meta))
    return _zip(root, out / "marapone-bid-leveling-template.zip")


def _bid_leveling_xlsx(path, rows, header, trades, universal, juris, meta):
    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="12161A")
    head_font = Font(bold=True, color="FFFFFF", size=9)
    accent = Font(bold=True, color="FF6A13", size=9)
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, ncols, freeze="A2"):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 26
        ws.freeze_panes = freeze

    # ── Tab 1: Scope Checklist ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Scope Checklist"
    ws.append([h.replace("_", " ").title() for h in header])
    for row in rows:
        ws.append([clean(c) for c in row])
    style_header(ws, len(header))
    for col, width in zip("ABCDEFGHIJ", [20, 11, 26, 11, 34, 11, 9, 12, 60, 40]):
        ws.column_dimensions[col].width = width
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=9).alignment = wrap
        ws.cell(row=r, column=10).alignment = wrap
        if ws.cell(row=r, column=7).value == "yes":
            ws.cell(row=r, column=7).font = accent

    # ── Tab 2: Comparison Matrix ─────────────────────────────────────────────
    cm = wb.create_sheet("Comparison Matrix")
    cm_header = (
        ["Element Key", "Trade Package", "CSI", "Scope Element", "Critical"]
        + [f"Bidder {chr(65 + i)} ($)" for i in range(BIDDER_COLS)]
        + ["Priced By", "Gap Flag", "Spread ($)", "Spread (%)", "Notes"]
    )
    cm.append(cm_header)
    first_bid, last_bid = 6, 5 + BIDDER_COLS       # columns F..H
    fb, lb = get_column_letter(first_bid), get_column_letter(last_bid)

    for i, row in enumerate(rows, start=2):
        cm.append([
            clean(row[0]), clean(row[2]), clean(row[3]), clean(row[4]), clean(row[6]),
            None, None, None,
            f"=COUNT({fb}{i}:{lb}{i})",
            (f'=IF(COUNT({fb}{i}:{lb}{i})=0,"NOT PRICED BY ANYONE",'
             f'IF(COUNT({fb}{i}:{lb}{i})<{BIDDER_COLS},'
             f'"GAP - "&({BIDDER_COLS}-COUNT({fb}{i}:{lb}{i}))&" of {BIDDER_COLS} missing",'
             f'""))'),
            f'=IF(COUNT({fb}{i}:{lb}{i})>1,MAX({fb}{i}:{lb}{i})-MIN({fb}{i}:{lb}{i}),"")',
            (f'=IF(AND(COUNT({fb}{i}:{lb}{i})>1,MIN({fb}{i}:{lb}{i})>0),'
             f'(MAX({fb}{i}:{lb}{i})-MIN({fb}{i}:{lb}{i}))/MIN({fb}{i}:{lb}{i}),"")'),
            None,
        ])
    style_header(cm, len(cm_header))
    for col, width in zip("ABCDEFGHIJKLM",
                          [20, 26, 11, 34, 9, 14, 14, 14, 10, 26, 13, 11, 34]):
        cm.column_dimensions[col].width = width
    for r in range(2, cm.max_row + 1):
        for c in range(first_bid, last_bid + 1):
            cm.cell(row=r, column=c).number_format = '#,##0.00'
        cm.cell(row=r, column=11).number_format = '#,##0.00'
        cm.cell(row=r, column=12).number_format = '0.0%'
        cm.cell(row=r, column=10).font = accent

    # Capture the last data row before writing the summary block: writing into a
    # cell moves max_row, so deriving the range afterwards silently drops rows.
    last_data = cm.max_row
    total = last_data + 2
    cm.cell(row=total, column=4, value="TOTAL QUOTED").font = Font(bold=True, size=9)
    cm.cell(row=total + 1, column=4, value="ELEMENTS NOT PRICED").font = Font(bold=True, size=9)
    for c in range(first_bid, last_bid + 1):
        letter = get_column_letter(c)
        cell = cm.cell(row=total, column=c, value=f"=SUM({letter}2:{letter}{last_data})")
        cell.font = Font(bold=True, size=9)
        cell.number_format = '#,##0.00'
        cm.cell(row=total + 1, column=c,
                value=f"=COUNTBLANK({letter}2:{letter}{last_data})").font = \
            Font(bold=True, color="FF6A13", size=9)

    # ── Tab 3: Region Notes ──────────────────────────────────────────────────
    rn = wb.create_sheet("Region Notes")
    rn.append(["Topic", "Ontario / Toronto position", "Value", "Source",
               "Why it changes a levelled number"])
    act = juris.get("construction_act", {})
    bond = juris.get("bonding_insurance", {})
    tax = juris.get("tax", {})
    pp = act.get("prompt_payment", {})
    region_rows = [
        ["Retainage terminology",
         "Called ‘holdback’ in Ontario, not ‘retainage’ or ‘retention’.",
         "—", clean(act.get("citation")),
         "A quote using ‘retention’ is often a template from another market. "
         "Check whether its payment terms were adjusted with it."],
        ["Statutory holdback", "Mandatory, not negotiable by contract.",
         pct(act.get("statutory_holdback_pct")), clean(act.get("citation")),
         "A sub who has not carried the financing cost of 10% held for "
         f'{clean(act.get("holdback_release_days"))}+ days is under-priced, not cheaper.'],
        ["Holdback release", "Days after substantial performance is published.",
         f'{clean(act.get("holdback_release_days"))} days', clean(act.get("citation")),
         clean(act.get("cost_impact"))],
        ["Prompt payment — owner to contractor", "From a proper invoice.",
         f'{clean(pp.get("owner_to_contractor_days"))} days', clean(act.get("citation")),
         clean(pp.get("note"))],
        ["Prompt payment — contractor to sub", "From receipt of owner payment.",
         f'{clean(pp.get("contractor_to_sub_days"))} days', clean(act.get("citation")),
         "Subcontract terms that say otherwise do not displace the Act. A quote "
         "priced on 60-day terms is priced on a fiction."],
        ["Lien preservation / perfection", "Statutory clocks.",
         f'{clean(act.get("lien_preservation_days"))} / '
         f'{clean(act.get("lien_perfection_days"))} days', clean(act.get("citation")),
         "Sets how long exposure to a sub's unpaid suppliers stays live."],
        ["Adjudication", clean(act.get("adjudication")), "—",
         clean(act.get("citation")),
         "Interim adjudication changes the leverage in a disputed change order."],
        ["Performance bond", "Percentage of contract value.",
         pct(bond.get("performance_bond_pct")), "Surety practice, Ontario",
         "A bonded and an unbonded quote are not comparable until the premium is added back."],
        ["Labour & material bond", "Percentage of contract value.",
         pct(bond.get("labour_material_bond_pct")), "Surety practice, Ontario",
         "Same — level bonded against bonded."],
        ["Bond premium", "Approximate cost of bonding.",
         pct(bond.get("bond_premium_pct_of_contract")), "Surety practice, Ontario",
         "The number to add to an unbonded quote before comparing it to a bonded one."],
        ["Tax treatment", "HST shown as a separate line, not buried in the rate.",
         pct(tax.get("hst_pct")), clean(tax.get("citation")),
         clean(tax.get("note"))],
    ]
    for row in region_rows:
        rn.append([clean(c) for c in row])
    style_header(rn, 5)
    for col, width in zip("ABCDE", [30, 44, 14, 40, 62]):
        rn.column_dimensions[col].width = width
    for r in range(2, rn.max_row + 1):
        for c in (2, 4, 5):
            rn.cell(row=r, column=c).alignment = wrap

    note_row = rn.max_row + 2
    rn.cell(row=note_row, column=1, value="COVERAGE").font = accent
    rn.cell(
        row=note_row, column=2,
        value=("These notes are Ontario/Toronto — the jurisdiction the Bid Leveler "
               "engine carries data for. The scope checklist itself is CSI-coded and "
               "structural: it levels a quote anywhere. What does not travel is this "
               "tab — holdback percentages, payment clocks and bonding norms are set "
               "by statute and differ by jurisdiction. Replace this tab with your "
               "own before levelling outside Ontario."),
    ).alignment = wrap
    rn.row_dimensions[note_row].height = 76

    # ── Tab 4: Trade Packages ────────────────────────────────────────────────
    tp = wb.create_sheet("Trade Packages")
    tp.append(["Trade Code", "Trade Package", "Short Name", "Scope Elements",
               "Also Called"])
    tp.append(["ALL", "Universal — applies to every trade", "Universal",
               len(universal), "Carve-outs common to all quotes"])
    for trade in trades:
        tp.append([clean(trade.get("code")), clean(trade.get("name")),
                   clean(trade.get("short")), len(trade.get("elements", [])),
                   clean(trade.get("aliases"))])
    style_header(tp, 5)
    for col, width in zip("ABCDE", [12, 40, 22, 15, 60]):
        tp.column_dimensions[col].width = width
    for r in range(2, tp.max_row + 1):
        tp.cell(row=r, column=5).alignment = wrap

    # ── Tab 5: Read Me ───────────────────────────────────────────────────────
    rm = wb.create_sheet("Read Me", 0)
    rm.column_dimensions["A"].width = 22
    rm.column_dimensions["B"].width = 104
    lines = [
        ("Marapone Bid-Levelling Template", ""),
        ("", ""),
        ("What this is",
         f"The scope-element checklist behind the Marapone Bid Leveler: "
         f"{len(rows)} elements across {len(trades)} trade packages, each CSI-coded. "
         f"{len(universal)} of them are universal — they apply to every trade and are "
         f"where most hidden cost lives."),
        ("How to use it",
         "1. Open 'Comparison Matrix'. 2. Delete the rows for trades you are not "
         "levelling. 3. Enter each bidder's amount per element in the Bidder columns. "
         "4. Leave a cell blank where a bidder did not price the element — blank is "
         "the signal, zero is not. 5. Read the Gap Flag and Spread columns."),
        ("Gap Flag",
         "Fires when fewer than all bidders priced an element. 'NOT PRICED BY ANYONE' "
         "usually means the element is missing from the tender documents, not that "
         "three subs independently forgot it."),
        ("Spread",
         "Absolute and percentage difference between the highest and lowest quote on "
         "that element. A large spread on one line is normally a scope "
         "misunderstanding, not a pricing advantage."),
        ("Critical elements",
         "Marked in the Critical column. These are the carve-outs that most often "
         "turn a low quote into a change order after award."),
        ("Coverage",
         "The scope checklist is CSI-coded and structural — it travels. The 'Region "
         "Notes' tab is Ontario/Toronto specific (holdback, prompt payment, bonding) "
         "and must be replaced before levelling in another jurisdiction."),
        ("Source",
         f"Exported {BUILD_DATE} from Marapone Bid Leveler trade_scopes.json "
         f"revision {clean(meta.get('revision'))}, region {clean(meta.get('region'))}."),
        ("Disclaimer", DISCLAIMER),
    ]
    for label, text in lines:
        rm.append([label, text])
    for r in range(2, rm.max_row + 1):
        rm.cell(row=r, column=1).font = Font(bold=True, size=9)
        rm.cell(row=r, column=2).alignment = wrap
        rm.row_dimensions[r].height = 42
    rm.cell(row=1, column=1).font = Font(bold=True, size=14, color="12161A")
    rm.row_dimensions[1].height = 22

    wb.save(str(path))


def _bid_how_to_pdf(path, rows, trades, universal, juris, meta):
    doc = PackDoc(path, "Bid-Levelling Template", "How to use it", wide=False)
    flow = title_block(
        "Bid-Levelling Template — how to use it",
        f"{len(rows)} scope elements across {len(trades)} trade packages, each "
        f"CSI-coded. {len(universal)} are universal: they apply to every trade and "
        "are the lines a sub can silently drop without the bid form ever showing it.",
        [
            ("Primary file", "marapone-bid-leveling-template.xlsx"),
            ("Tabs", "Read Me, Scope Checklist, Comparison Matrix, Region Notes, "
                     "Trade Packages"),
            ("Also included", "scope-checklist.csv — the same rows, plain CSV"),
            ("Scope checklist coverage", "CSI-coded and structural — travels"),
            ("Region notes coverage", "Ontario / Toronto — does not travel"),
            ("Exported", BUILD_DATE),
        ],
    )
    flow += [
        para("The method", "h2"),
        para(
            "Levelling is not comparing three totals. It is establishing that three "
            "subcontractors priced the <i>same</i> work, and pricing the difference "
            "where they did not. The checklist is the common denominator: every "
            "element one of them might have excluded, in one list, so an exclusion "
            "shows up as a blank cell instead of as a change order in month four.",
        ),
        para("Step by step", "h3"),
        para(
            "<b>1.</b> Open the <b>Comparison Matrix</b> tab.<br/>"
            "<b>2.</b> Delete rows for trades you are not levelling.<br/>"
            "<b>3.</b> Enter each bidder's amount against each element.<br/>"
            "<b>4.</b> Leave the cell <b>blank</b> where a bidder did not price it. "
            "Blank is the signal; a zero says ‘priced at nil’ and is a different "
            "claim.<br/>"
            "<b>5.</b> Work the <b>Gap Flag</b> column before the totals. A gap is "
            "worth more than a spread.",
        ),
        para("Reading the flags", "h3"),
        para(
            "<b>GAP — n of 3 missing.</b> Some bidders carried the element and "
            "others did not. Whatever the low bidder saved here, you will pay later.<br/>"
            "<b>NOT PRICED BY ANYONE.</b> Usually the tender documents are silent on "
            "it, not that three subs independently forgot. Issue an addendum.<br/>"
            "<b>Large spread on one line.</b> Almost always a scope "
            "misunderstanding rather than a genuine pricing advantage. Ask both.",
        ),
        para("The universal elements", "h2"),
        para(
            "These apply to every trade regardless of package, and they are where "
            "the money hides — a quote can be complete on its own trade and still "
            "leave out the hoisting that gets its material to the fourth floor.",
        ),
        data_table(
            ["Element", "CSI", "Critical", "What to look for"],
            [[u.get("label"), u.get("csi"), "yes" if u.get("critical") else "",
              u.get("description")] for u in universal],
            [1.35 * inch, 0.62 * inch, 0.5 * inch, None],
        ),
        PageBreak(),
        para("Trade packages", "h2"),
        data_table(
            ["Code", "Trade package", "Elements", "Also called"],
            [["ALL", "Universal — applies to every trade", str(len(universal)),
              "Carve-outs common to all quotes"]]
            + [[t.get("code"), t.get("name"), str(len(t.get("elements", []))),
                t.get("aliases")] for t in trades],
            [0.5 * inch, 1.9 * inch, 0.6 * inch, None],
        ),
        para("Coverage", "h2"),
        para(
            "The scope checklist is CSI-coded and structural. What a mechanical "
            "subcontractor can leave out of a quote does not change at a border, so "
            "it levels a quote anywhere.",
        ),
        para(
            "The <b>Region Notes</b> tab is the part that does not travel. Holdback "
            "percentages, prompt-payment clocks and bonding norms are set by "
            "statute, and the ones shipped here are Ontario's, from the "
            "<i>Construction Act</i>. Levelling outside Ontario: keep the checklist, "
            "replace that tab.",
        ),
    ]
    doc.build(flow)


def _bid_readme(rows, trades, universal, meta) -> str:
    return f"""# Bid-Levelling Template

The scope-element checklist behind the
[Bid Leveler](https://maraponecontracting.com/construction/bidleveler):
**{len(rows)} scope elements** across **{len(trades)} trade packages**, each CSI-coded.

{len(universal)} of those elements are *universal* — they apply to every trade and
are where most hidden cost lives.

Exported {BUILD_DATE} from Bid Leveler `trade_scopes.json` revision
{clean(meta.get('revision'))}.

## Files

```
marapone-bid-leveling-template.xlsx   primary file, 5 tabs
scope-checklist.csv                   the same {len(rows)} rows, plain CSV
how-to-use.pdf                        the method, written out
```

### Tabs in the workbook

| Tab | What it is |
|---|---|
| Read Me | Method and coverage, in the file itself |
| Scope Checklist | All {len(rows)} elements, with what to look for in a quote |
| Comparison Matrix | Blank, with live formulas for {BIDDER_COLS} bidders — gap flags, spread, totals |
| Region Notes | Ontario/Toronto statutory positions that change a levelled number |
| Trade Packages | The {len(trades)} packages and their element counts |

## Using the Comparison Matrix

Enter each bidder's amount per element. **Leave a cell blank where a bidder did
not price the element** — blank is the signal, zero is a different claim
("priced at nil"). The Gap Flag and Spread columns compute themselves.

A gap is worth more than a spread. `NOT PRICED BY ANYONE` usually means the
tender documents are silent, not that three subs independently forgot.

## Coverage

The **scope checklist travels**. It is CSI-coded and structural: what a
mechanical subcontractor can leave out of a quote does not change at a border.

The **Region Notes tab does not travel**. Holdback percentages, prompt-payment
clocks and bonding norms are set by statute, and the ones here are Ontario's,
from the *Construction Act*. Levelling outside Ontario: keep the checklist,
replace that tab.

## Disclaimer

Reference material only — not a substitute for review by a licensed
professional in your jurisdiction.
"""


# ── Pack 5 — Master Scope Requirement Checklist ──────────────────────────────


def build_pack_scope_checklist(engines: Path, out: Path) -> Path:
    knowledge = engines / "ScopeGuard" / "resources" / "knowledge"
    data = load(knowledge / "trade_scopes.json")
    intel = load(knowledge / "ontario_intel.json")
    interfaces = load(knowledge / "interface_matrix.json")

    root = out / "build" / "pack5-master-scope-checklist"
    if root.exists():
        shutil.rmtree(root)
    root.mkdir(parents=True)

    meta = data.get("_meta", {})
    trades = data.get("trades", [])

    rows: list[list] = []
    for trade in trades:
        for req in trade.get("requirements", []):
            rows.append([
                req.get("code"), trade.get("code"), trade.get("name"),
                req.get("csi") or clean(trade.get("csi")), req.get("title"),
                req.get("criticality"), req.get("unit"), req.get("basis"),
                req.get("spec_ref"), req.get("keywords"),
                pct(trade.get("typical_share")),
                "Marapone ScopeGuard trade_scopes.json", clean(meta.get("revision")),
            ])

    write_csv(
        root / "marapone-master-scope-checklist.csv",
        ["requirement_code", "trade_code", "trade_package", "csi_code",
         "requirement", "criticality", "unit", "quantity_basis", "spec_reference",
         "keywords", "trade_typical_share_of_contract", "source", "last_checked"],
        rows,
    )

    _scope_checklist_pdf(root / "marapone-master-scope-checklist.pdf", trades, rows, meta)
    _scope_region_notes_pdf(root / "region-notes.pdf", trades, intel, interfaces)
    write_text(root / "README.md", _scope_readme(trades, rows, meta))
    return _zip(root, out / "marapone-master-scope-checklist.zip")


def _scope_checklist_pdf(path, trades, rows, meta):
    doc = PackDoc(path, "Master Scope Requirement Checklist",
                  f"{len(rows)} requirements · {len(trades)} trade packages",
                  wide=True)
    flow = title_block(
        "Master Scope Requirement Checklist",
        f"{len(rows)} scope requirements across {len(trades)} trade packages, each "
        "carrying its CSI code and the specification section it is normally "
        "written under. This is the library ScopeGuard checks a subcontract scope "
        "letter against.",
        [
            ("Requirements", str(len(rows))),
            ("Trade packages", str(len(trades))),
            ("Coded to", "CSI MasterFormat"),
            ("Source revision", clean(meta.get("revision"))),
            ("Source region", clean(meta.get("region"))),
            ("Exported", BUILD_DATE),
            ("Machine-readable", "marapone-master-scope-checklist.csv"),
        ],
    )
    flow += [
        para("How to use it", "h3"),
        para(
            "Read down the trade you are awarding and confirm each requirement is "
            "either in the scope letter or deliberately excluded from it. A "
            "requirement that is in neither is the one that becomes a change order. "
            "<b>Criticality</b> ranks how often that happens and how much it costs "
            "when it does.", "small",
        ),
        para(
            "<b>Unit</b> and <b>quantity basis</b> are what the requirement is "
            "normally measured in — useful when a scope letter says ‘as required’ "
            "and you need to make it a number.", "small",
        ),
        PageBreak(),
    ]

    by_trade: dict[str, list[list]] = {}
    for row in rows:
        by_trade.setdefault(clean(row[2]), []).append(row)

    widths = [0.72 * inch, 3.0 * inch, 0.78 * inch, 0.72 * inch, 0.52 * inch,
              1.05 * inch, 2.55 * inch]
    for trade in trades:
        name = clean(trade.get("name"))
        group = by_trade.get(name, [])
        if not group:
            continue
        flow.append(para(
            f'{trade.get("code")} — {name} · {len(group)} requirements '
            f'· typically {pct(trade.get("typical_share"))} of contract value',
            "h2",
        ))
        flow.append(data_table(
            ["Code", "Requirement", "CSI", "Critical", "Unit", "Basis",
             "Spec section"],
            [[r[0], r[4], r[3], r[5], r[6], r[7], r[8]] for r in group],
            widths,
        ))
    doc.build(flow)


def _scope_region_notes_pdf(path, trades, intel, interfaces):
    meta = intel.get("_meta", {})
    doc = PackDoc(path, "Master Scope Checklist — Region Notes",
                  "Ontario / Toronto")
    flow = title_block(
        "Region Notes — Ontario and Toronto",
        "The checklist itself is CSI-coded and structural. These are the things "
        "that change what a requirement <i>means</i> in this jurisdiction: statutory "
        "clauses that override a subcontract, contract forms that behave "
        "differently, and the disputes that recur.",
        [
            ("Jurisdiction", clean(meta.get("jurisdiction"))),
            ("Source revision", clean(meta.get("revision"))),
            ("Reviewed on", clean(meta.get("reviewed_on"))),
            ("Exported", BUILD_DATE),
            ("Applies to", "The interpretation layer only — not the requirement list"),
        ],
    )
    flow += [
        para(
            "<b>Read this before applying the checklist outside Ontario.</b> The "
            "requirements travel; the statutory positions below do not. Holdback, "
            "payment clocks and lien periods are set by provincial statute and "
            "differ in every jurisdiction.", "small",
        ),
    ]

    for section, node in intel.items():
        if section == "_meta" or not isinstance(node, dict):
            continue
        flow.append(para(humanise(section), "h2"))
        for key, value in node.items():
            if isinstance(value, dict):
                flow.append(para(humanise(key), "h3"))
                rows = [[humanise(k),
                         pct(v) if isinstance(v, float) and "pct" in k else clean(v)]
                        for k, v in value.items()]
                flow.append(data_table(["Item", "Detail"], rows, [1.7 * inch, None]))
            elif isinstance(value, list):
                flow.append(para(humanise(key), "h3"))
                for item in value:
                    if isinstance(item, dict):
                        rows = [[humanise(k), clean(v)]
                                for k, v in item.items()]
                        flow.append(data_table(["Item", "Detail"], rows,
                                               [1.7 * inch, None]))
                        flow.append(Spacer(1, 4))
                    else:
                        flow.append(para(f"• {clean(item)}", "small"))
            else:
                flow.append(para(f'<b>{humanise(key)}.</b> '
                                 f'{clean(value)}', "small"))

    trade_notes = [(clean(t.get("name")), clean(t.get("toronto_note")))
                   for t in trades if t.get("toronto_note")]
    if trade_notes:
        flow += [
            PageBreak(),
            para("Per-trade notes — Toronto", "h2"),
            para(
                "What changes about this trade's scope specifically on a Toronto "
                "project. These are the local conditions that turn a standard "
                "requirement into a priced one.", "small",
            ),
            data_table(["Trade package", "Toronto note"], trade_notes,
                       [1.65 * inch, None]),
        ]

    rows = interfaces.get("interfaces", [])
    if rows:
        flow += [
            PageBreak(),
            para(f"Trade interfaces — {len(rows)} recurring gaps", "h2"),
            para(
                "Scope gaps do not usually sit inside a trade; they sit between "
                "two. Each row is a boundary where both parties routinely assume "
                "the other carried it.", "small",
            ),
            data_table(
                ["Interface", "Detail"],
                [[clean(i.get("name") or i.get("key") or i.get("id")),
                  "; ".join(f"{k}: {clean(v)}" for k, v in i.items()
                            if k not in ("name", "key", "id"))]
                 for i in rows],
                [1.65 * inch, None],
            ),
        ]
    doc.build(flow)


def _scope_readme(trades, rows, meta) -> str:
    table = "\n".join(
        f"| {clean(t.get('code'))} | {clean(t.get('name'))} | "
        f"{len(t.get('requirements', []))} | {pct(t.get('typical_share'))} |"
        for t in trades
    )
    return f"""# Master Scope Requirement Checklist

The scope-requirement library behind
[ScopeGuard](https://maraponecontracting.com/construction/scopeguard):
**{len(rows)} requirements** across **{len(trades)} trade packages**, each carrying
its CSI code and the specification section it is normally written under.

Exported {BUILD_DATE} from ScopeGuard `trade_scopes.json` revision
{clean(meta.get('revision'))}.

## Files

```
marapone-master-scope-checklist.csv   {len(rows)} requirements, one row each
marapone-master-scope-checklist.pdf   the same rows, grouped by trade
region-notes.pdf                      Ontario/Toronto interpretation layer
```

## How to use it

Read down the trade you are awarding and confirm each requirement is either in
the scope letter or deliberately excluded from it. A requirement in neither is
the one that becomes a change order.

`criticality` ranks how often that happens and what it costs. `unit` and
`quantity_basis` are what the requirement is normally measured in — useful when
a scope letter says "as required" and you need to turn it into a number.

## Trade packages

| Code | Trade package | Requirements | Typical share of contract |
|---|---|---|---|
{table}

## Coverage

The **requirement list travels**. It is CSI-coded and structural — what a
drywall subcontractor is expected to carry does not change at a border.

`region-notes.pdf` is the part that does not. It holds the Ontario statutory
layer (Construction Act holdback, prompt payment, lien clocks), Toronto
per-trade conditions, and the recurring trade-interface gaps. Holdback,
payment clocks and lien periods are set by provincial statute and differ
everywhere. Applying the checklist outside Ontario: keep the requirements,
replace the region notes.

## Disclaimer

Reference material only — not a substitute for review by a licensed
professional in your jurisdiction.
"""


# ── Bundle ───────────────────────────────────────────────────────────────────

BUNDLE_FOLDERS = [
    ("pack1-building-code-compliance-library", "1-building-code-compliance-library",
     "Building Code Compliance Library", "$89"),
    ("pack2-assembly-property-schema", "2-assembly-property-schema",
     "Assembly Property Schema", "$89"),
    ("pack3-estimating-rate-pack-2026", "3-estimating-rate-pack-2026",
     "Estimating Rate Pack — 2026 Edition", "$99"),
    ("pack4-bid-leveling-template", "4-bid-leveling-template",
     "Bid-Levelling Template", "$59"),
    ("pack5-master-scope-checklist", "5-master-scope-checklist",
     "Master Scope Requirement Checklist", "$59"),
]


def build_bundle(out: Path) -> Path:
    root = out / "build" / "bundle"
    if root.exists():
        shutil.rmtree(root)
    root.mkdir(parents=True)

    summary = []
    for src_name, dest_name, label, price in BUNDLE_FOLDERS:
        src = out / "build" / src_name
        shutil.copytree(src, root / dest_name)
        files = sum(1 for p in (root / dest_name).rglob("*") if p.is_file())
        summary.append((dest_name, label, price, files))

    _bundle_index_pdf(root / "START-HERE.pdf", summary)
    write_text(root / "README.md", _bundle_readme(summary))
    return _zip(root, out / "marapone-full-data-pack-bundle.zip")


def _bundle_index_pdf(path, summary):
    doc = PackDoc(path, "The Full Data Pack Bundle", "Start here")
    flow = title_block(
        "The Full Data Pack Bundle",
        "All five Marapone reference libraries, unmodified, in one download. Each "
        "folder is exactly the pack sold on its own — same files, same README, "
        "same schema — so anything written against a single pack keeps working.",
        [
            ("Packs", str(len(summary))),
            ("Files", str(sum(s[3] for s in summary))),
            ("Bought separately", "$395"),
            ("Bundle", "$349"),
            ("Exported", BUILD_DATE),
        ],
    )
    flow += [
        para("What is in each folder", "h2"),
        data_table(
            ["Folder", "Pack", "Sold at", "Files"],
            [[s[0], s[1], s[2], str(s[3])] for s in summary],
            [2.05 * inch, 2.35 * inch, 0.7 * inch, 0.5 * inch],
        ),
        para(
            "Every folder carries its own README.md with that pack's coverage, "
            "schema and sources. Start there for anything pack-specific.", "small",
        ),
        para("How the five fit together", "h2"),
        para(
            "They share a spine: <b>CSI MasterFormat</b>. A scope requirement in "
            "pack 5 carries a CSI code; the assembly it describes is classified "
            "under the same code in pack 2; the scope element you level a quote "
            "against in pack 4 uses it too. That is what lets a requirement, its "
            "assembly and its rate line up across packs rather than needing to be "
            "reconciled by hand.",
        ),
        para("Coverage, stated once", "h2"),
        data_table(
            ["Pack", "Travels", "Jurisdiction-bound"],
            [
                ["1 · Building Code", "Category structure only",
                 "Ontario + City of Toronto — clause numbers and dimensions"],
                ["2 · Assembly Schema",
                 "Properties, CSI division/section spine (North America)",
                 "Some designation values are CSA/ANSI"],
                ["3 · Estimating Rates", "The contingency and risk framework",
                 "Ontario statutory + City of Toronto municipal figures"],
                ["4 · Bid Levelling", "The {n}-element scope checklist".format(n=155),
                 "Region Notes tab only — Ontario holdback, payment, bonding"],
                ["5 · Scope Checklist", "The 215 requirements, CSI-coded",
                 "region-notes.pdf only — Ontario statutory, Toronto conditions"],
            ],
            [1.25 * inch, 2.0 * inch, None],
        ),
        para(
            "Packs 2, 4 and 5 are standards-based and usable anywhere. Packs 1 and "
            "3 are statutory and municipal: they are Ontario and Toronto, and this "
            "edition does not carry other provinces, US states or EU jurisdictions.",
            "small",
        ),
        para("Licence", "h2"),
        para(
            "One-time purchase. Use these files in your own projects, spreadsheets, "
            "estimates and internal tools without limit. Do not resell or "
            "redistribute them as a data product.",
        ),
    ]
    doc.build(flow)


def _bundle_readme(summary) -> str:
    table = "\n".join(f"| `{s[0]}/` | {s[1]} | {s[2]} | {s[3]} |" for s in summary)
    return f"""# The Full Data Pack Bundle

All five Marapone reference libraries, unmodified, in one download.
Exported {BUILD_DATE}.

| Folder | Pack | Sold separately at | Files |
|---|---|---|---|
{table}

Each folder is exactly the pack sold on its own — same files, same README, same
schema. Anything written against a single pack keeps working here.

## How the five fit together

They share a spine: **CSI MasterFormat**. A scope requirement in pack 5 carries
a CSI code; the assembly it describes is classified under the same code in pack
2; the scope element you level a quote against in pack 4 uses it too.

## Coverage

Packs **2, 4 and 5** are standards-based and travel. Packs **1 and 3** are
statutory and municipal — **Ontario and the City of Toronto**. This edition
does not carry other provinces, US states or EU jurisdictions. Each pack's own
README states its coverage precisely.

## Licence

One-time purchase. Use these files in your own projects, spreadsheets,
estimates and internal tools without limit. Do not resell or redistribute them
as a data product.

## Disclaimer

Reference material only — not a substitute for review by a licensed
professional in your jurisdiction.
"""


# ── Packaging + QA ───────────────────────────────────────────────────────────


def _zip(folder: Path, target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists():
        target.unlink()
    with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
        for path in sorted(folder.rglob("*")):
            if path.is_file():
                zf.write(path, path.relative_to(folder))
    return target


def qa(artifacts: list[Path]) -> bool:
    """Open every zip, prove it is readable, and refuse to pass on placeholder text."""
    banned = ("lorem ipsum", "tbd", "todo", "xxx", "placeholder", "coming soon",
              "sample text", "fixme")
    ok = True
    print("\n── QA ──")
    for artifact in artifacts:
        with zipfile.ZipFile(artifact) as zf:
            bad = zf.testzip()
            names = zf.namelist()
            if bad:
                print(f"  FAIL  {artifact.name}: corrupt member {bad}")
                ok = False
                continue
            empty = [n for n in names if not n.endswith("/") and zf.getinfo(n).file_size == 0]
            if empty:
                print(f"  FAIL  {artifact.name}: empty files {empty}")
                ok = False
            hits = []
            for name in names:
                if name.lower().endswith((".csv", ".md", ".json")):
                    text = zf.read(name).decode("utf-8-sig", "replace").lower()
                    for word in banned:
                        if word in text:
                            hits.append(f"{name}:{word}")
            if hits:
                print(f"  FAIL  {artifact.name}: placeholder text {hits[:4]}")
                ok = False
            size = artifact.stat().st_size / 1024
            print(f"  pass  {artifact.name:52s} {len(names):4d} files  {size:8.1f} KB")
    return ok


# ── Entry point ──────────────────────────────────────────────────────────────


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--engines", type=Path, default=DEFAULT_ENGINES,
                    help="path to the MaraponeAI-Tools checkout")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT,
                    help="where to write the packs")
    args = ap.parse_args()

    engines, out = args.engines.resolve(), args.out.resolve()
    if not engines.exists():
        print(f"Engine data not found at {engines}", file=sys.stderr)
        return 1
    out.mkdir(parents=True, exist_ok=True)

    print(f"Engines : {engines}")
    print(f"Output  : {out}\n")

    artifacts = []
    for label, fn in [
        ("Pack 1  Building Code Compliance Library", build_pack_code_library),
        ("Pack 2  Assembly Property Schema", build_pack_assembly_schema),
        ("Pack 3  Estimating Rate Pack 2026", build_pack_rate),
        ("Pack 4  Bid-Levelling Template", build_pack_bid_leveling),
        ("Pack 5  Master Scope Requirement Checklist", build_pack_scope_checklist),
    ]:
        print(f"  building  {label}")
        artifacts.append(fn(engines, out))

    print("  building  Bundle  The Full Data Pack Bundle")
    artifacts.append(build_bundle(out))

    if not qa(artifacts):
        print("\nQA failed — not fit to sell.", file=sys.stderr)
        return 1

    print(f"\n{len(artifacts)} SKUs built in {out}")
    print("Upload each to private storage and set the matching PACK_URL_* env var")
    print("(see lib/fulfillment.js) to switch that SKU to instant delivery.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
